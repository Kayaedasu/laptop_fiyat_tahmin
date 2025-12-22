import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("\n📊 ADIM 4: Excel Raporu Oluşturuluyor...")

df = pd.read_csv("veriler/birlesik/laptops_sayisal_donusum.csv", sep=';')

wb = Workbook()
ws = wb.active
ws.title = "Laptop Verileri"

# Tasarım Ayarları
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2b5797') # Kurumsal Mavi
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Başlıkları Yaz
headers = df.columns.tolist()
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Verileri Yaz
for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        
        # Formatlama
        col_name = headers[col_idx-1]
        if 'Fiyat' in col_name:
            cell.number_format = '#,##0.00 ₺'
        elif 'GB' in col_name:
            cell.alignment = Alignment(horizontal='right')

# Filtre ve Bölmeleri Dondur
ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(df)+1}"
ws.freeze_panes = 'A2'

# Genişlik Ayarları
ws.column_dimensions['A'].width = 15  # Marka
ws.column_dimensions['B'].width = 30  # Model
ws.column_dimensions['C'].width = 25  # İşlemci
ws.column_dimensions['F'].width = 25  # Ekran Kartı

wb.save("veriler/birlesik/laptops_rapor.xlsx")
print(f"🎉 Rapor başarıyla oluşturuldu: veriler/birlesik/laptops_rapor.xlsx")

# Konsola küçük bir özet
print("\n--- ÖZET ---")
print(f"Toplam Laptop: {len(df)}")
print(f"Marka Sayısı: {df['Marka'].nunique()}")
print(f"Ortalama Fiyat: {df['Fiyat'].mean():.2f} ₺")